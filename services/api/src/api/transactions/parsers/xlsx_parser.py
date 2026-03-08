from collections import Counter
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from api.models.bank_profile import BankProfile
from api.models.transaction import TransactionDirection
from api.transactions.parsers.base import ParsedTransaction, compute_dedup_hash


async def parse_xlsx(content: bytes, profile: BankProfile) -> list[ParsedTransaction]:
  """Parse XLSX content using the column mapping from a BankProfile."""
  wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
  ws = wb.active
  if ws is None:
    return []

  col = profile.column_map
  date_col_name = col["transaction_date"]
  amount_col_name = col["amount"]
  description_col_name = col.get("description", "")
  account_number = ""

  # Extract account number from metadata rows if present
  account_col_name = col.get("account_number_cell")
  if account_col_name:
    # e.g. "B1" -> row 1, col B
    for row in ws.iter_rows(min_row=1, max_row=profile.skip_rows, values_only=False):
      for cell in row:
        if cell.coordinate == account_col_name:
          account_number = str(cell.value or "").strip()

  # Read header row to find column indices
  header_row_idx = profile.skip_rows + 1
  headers: list[str] = []
  for row in ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True):
    headers = [str(c or "") for c in row]
    break

  if not headers:
    wb.close()
    return []

  # Map column names to indices
  col_indices: dict[str, int] = {}
  for i, h in enumerate(headers):
    col_indices[h] = i

  date_idx = col_indices.get(date_col_name)
  amount_idx = col_indices.get(amount_col_name)
  desc_idx = col_indices.get(description_col_name) if description_col_name else None

  if date_idx is None or amount_idx is None:
    wb.close()
    return []

  balance_col_name = col.get("balance")
  hash_counts: Counter[str] = Counter()

  transactions: list[ParsedTransaction] = []
  for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
    cells = list(row)
    date_val = cells[date_idx] if date_idx < len(cells) else None
    amount_val = cells[amount_idx] if amount_idx < len(cells) else None
    desc_val = cells[desc_idx] if desc_idx is not None and desc_idx < len(cells) else None

    if date_val is None or amount_val is None:
      continue

    # Parse date
    if isinstance(date_val, date):
      tx_date = date_val
    else:
      date_str = str(date_val).strip()
      if not date_str:
        continue
      tx_date = date.fromisoformat(date_str)

    # Parse amount
    if isinstance(amount_val, (int, float)):
      amount_float = float(amount_val)
    else:
      amount_str = str(amount_val).strip()
      if not amount_str:
        continue
      amount_normalized = amount_str.replace(profile.decimal_separator, ".") if profile.decimal_separator != "." else amount_str
      amount_float = float(amount_normalized)

    if amount_float < 0:
      direction = TransactionDirection.debit
      amount_minor = round(abs(amount_float) * 100)
    else:
      direction = TransactionDirection.credit
      amount_minor = round(amount_float * 100)

    if amount_minor == 0:
      continue

    description = str(desc_val or "").strip()
    date_str_for_hash = tx_date.isoformat()
    amount_str_for_hash = str(amount_val)

    # Build dedup description from balance + description when available.
    balance_str = ""
    if balance_col_name:
      balance_idx = col_indices.get(balance_col_name)
      if balance_idx is not None and balance_idx < len(cells):
        balance_str = str(cells[balance_idx] or "")

    dedup_desc = f"{balance_str}:{description}" if balance_str else description

    base_hash = compute_dedup_hash(
      bank=profile.bank_name,
      account_number=account_number,
      date=date_str_for_hash,
      amount=amount_str_for_hash,
      description=dedup_desc,
    )

    # Append occurrence index so truly identical rows get unique hashes
    occurrence = hash_counts[base_hash]
    hash_counts[base_hash] += 1
    dedup_hash = f"{base_hash}:{occurrence}" if occurrence > 0 else base_hash

    raw_data = {headers[i]: str(v) if v is not None else "" for i, v in enumerate(cells) if i < len(headers)}

    transactions.append(
      ParsedTransaction(
        transaction_date=tx_date,
        amount_minor=amount_minor,
        currency=profile.currency,
        direction=direction,
        dedup_hash=dedup_hash,
        description=description,
        raw_data=raw_data,
      )
    )

  wb.close()
  return transactions
